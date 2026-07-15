"""Per-format text extractors for read_attachment_text (untrusted input).

The public handler lives in attachment_text.py; THIS module owns the
pure `bytes -> str` extractors it composes, one per supported format
(PDF, XLSX, CSV, plain text), plus the safe-wrap that turns any parser
failure into a typed signal instead of an exception escaping the read.

Threat model
------------
Every function here parses ATTACKER-CONTROLLED bytes: a PDF or XLSX
from any sender, fetched by the download path and handed in raw. PDF
and XLSX parsers have a CVE history (malformed streams, deeply nested
objects, zip bombs), so each extractor is bounded hard:

- SAFE-WRAP: every parser call is wrapped so RecursionError AND any
  broad Exception degrade to `ExtractionError` (a module-local
  sentinel the handler maps to a typed extraction_failed bad_request).
  The parsers NEVER raise past this layer to a -32603 / 500, and the
  bounds below stop them hanging unboundedly. Only a generic reason is
  surfaced; the file bytes/content are NEVER logged or echoed.
- EARLY-EXIT AT THE CAP: extraction stops the moment accumulated text
  reaches `max_chars`. A malicious huge / decompression-bomb file is
  never fully parsed-then-truncated (that would defeat the memory/time
  bound); we stop joining PDF pages / streaming XLSX cells / reading
  CSV rows as soon as the cap is hit.
- ITERATION GUARDS (belt-and-suspenders): a hard page cap (PDF), a
  hard row cap (XLSX / CSV), so a crafted file that streams unboundedly
  WITHOUT tripping the char cap (e.g. millions of empty rows) still
  terminates.

The library choices (pypdf, openpyxl) do NO network fetch, NO external
entity resolution, and NO code execution on parse. openpyxl runs in
`read_only=True, data_only=True` streaming mode so a zip bomb is bounded
by the streaming cursor plus the caps here. There is NO local byte-size
guard in this module: input size is bounded by Gmail's own ~25 MiB
attachment limit (external, not enforced here), and CPU/memory are
bounded by the streaming read_only mode plus the row/cell/char caps
below. pypdf's base install is pure-Python text extraction with no
image/crypto extras pulled.
"""

from __future__ import annotations

import csv
import io
import logging

logger = logging.getLogger(__name__)


# Belt-and-suspenders iteration guards. These bound the NUMBER of units
# processed even when the per-unit text is tiny, so a file that streams
# unboundedly (millions of near-empty PDF pages / XLSX or CSV rows)
# without ever hitting the char cap still terminates. Generous relative
# to any real invoice/order document, tight relative to a DoS payload.
_MAX_PDF_PAGES = 500
_MAX_SHEET_ROWS = 100_000
_MAX_CSV_ROWS = 100_000

# Total-cell budget for XLSX extraction. The row cap alone does not bound
# CPU when rows are WIDE: a crafted all-tiny-cell sheet (few rows, tens of
# thousands of columns each, or many rows whose cells render empty so they
# never accumulate toward the char cap) can ride the row cap and still
# make openpyxl parse a very large number of cell elements per call. A
# cell budget bounds the ACTUAL work regardless of the row/column shape,
# which a pure row cap cannot. Checked inside the iter_rows loop; when the
# budget is reached extraction aborts and returns what it has with the
# truncation flag set. Sized well above any real invoice/export (a few
# thousand rows of a few dozen columns is far under this) yet tight enough
# that a bomb terminates in well under a second.
_XLSX_MAX_CELLS = 200_000


class ExtractionError(Exception):
    """A parser failed on (potentially hostile) input.

    Module-local sentinel, NOT a caller-facing shape. The handler in
    attachment_text.py maps it to a typed extraction_failed
    bad_request. Carries a GENERIC reason only; the file bytes/content
    are never attached to it or logged.
    """


def _safe(reason: str) -> ExtractionError:
    """Log a generic reason (never file content) and build ExtractionError."""
    logger.warning("attachment text extraction failed: %s", reason)
    return ExtractionError(reason)


def extract_pdf_text(data: bytes, *, max_chars: int) -> str:
    """Extract text from PDF bytes, bounded by max_chars and a page cap.

    Joins per-page `extract_text()` output in page order, stopping as
    soon as the accumulated length reaches `max_chars` (early-exit: a
    huge PDF is never fully walked) or the `_MAX_PDF_PAGES` guard is
    hit. Any pypdf failure (malformed/hostile file) OR RecursionError
    (deeply nested object tree) degrades to ExtractionError. The caller
    applies the real cap+marker; returning slightly over `max_chars`
    here is fine (the handler truncates).
    """
    # Import inside the function so a pypdf import error cannot break
    # module import for the CSV/text paths, and so the dependency is
    # only touched when a PDF is actually parsed.
    from pypdf import PdfReader
    from pypdf.errors import PyPdfError

    try:
        reader = PdfReader(io.BytesIO(data))
        parts: list[str] = []
        total = 0
        for page_num, page in enumerate(reader.pages):
            if page_num >= _MAX_PDF_PAGES or total >= max_chars:
                break
            page_text = page.extract_text() or ""
            parts.append(page_text)
            total += len(page_text)
        return "\n".join(parts)
    except RecursionError as exc:
        raise _safe("pdf parse recursion limit") from exc
    except PyPdfError as exc:
        raise _safe("pdf parse error") from exc
    except Exception as exc:  # noqa: BLE001 - hostile PDF must not crash the read
        raise _safe("pdf extraction failed") from exc


def extract_xlsx_text(data: bytes, *, max_chars: int) -> tuple[str, bool]:
    """Extract cell values from XLSX bytes, bounded by max_chars, row cap, and cell budget.

    Streams every sheet in `read_only=True, data_only=True` mode (so a
    zip bomb is bounded by the streaming cursor rather than a full
    in-memory expand), rendering each row's non-empty cell values as a
    tab-joined line. Stops as soon as accumulated text reaches
    `max_chars` (early-exit), `_MAX_SHEET_ROWS` rows have been emitted, or
    the `_XLSX_MAX_CELLS` total-cell budget is reached across all sheets.
    The cell budget is the tighter guard for WIDE sheets: the row cap
    cannot bound a few rows of tens of thousands of columns, or many rows
    whose cells render empty (never accumulating toward the char cap), so
    the cell count bounds the actual openpyxl work regardless of shape.
    `data_only=True` returns the last-cached computed value for formula
    cells instead of the formula string.

    Returns `(text, budget_truncated)` where `budget_truncated` is True
    when extraction stopped because the cell budget was reached while more
    rows remained (so the caller marks the result truncated even though
    the char cap may not have fired). The char-cap and row-cap early-exits
    are surfaced through the normal char-cap truncation path in the
    handler and do not set this flag. Any openpyxl failure OR
    RecursionError degrades to ExtractionError. The workbook is always
    closed (read_only mode holds a file handle).
    """
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    wb = None
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        lines: list[str] = []
        total = 0
        rows_emitted = 0
        cells_seen = 0
        budget_truncated = False
        for sheet in wb.worksheets:
            if total >= max_chars or rows_emitted >= _MAX_SHEET_ROWS:
                break
            if cells_seen >= _XLSX_MAX_CELLS:
                budget_truncated = True
                break
            lines.append(f"# {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                if total >= max_chars or rows_emitted >= _MAX_SHEET_ROWS:
                    break
                if cells_seen >= _XLSX_MAX_CELLS:
                    budget_truncated = True
                    break
                # Count every cell this row costs BEFORE rendering, so the
                # budget bounds openpyxl's per-cell work regardless of how
                # many cells render to non-empty text.
                cells_seen += len(row)
                cells = ["" if v is None else str(v) for v in row]
                line = "\t".join(cells).rstrip("\t")
                if line:
                    lines.append(line)
                    total += len(line)
                rows_emitted += 1
        return "\n".join(lines), budget_truncated
    except RecursionError as exc:
        raise _safe("xlsx parse recursion limit") from exc
    except InvalidFileException as exc:
        raise _safe("xlsx parse error") from exc
    except Exception as exc:  # noqa: BLE001 - hostile XLSX must not crash the read
        raise _safe("xlsx extraction failed") from exc
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:  # noqa: BLE001 - close failure must not mask the result
                pass


def extract_csv_text(text: str, *, max_chars: int) -> str:
    """Render already-decoded CSV text into readable tab-joined rows.

    Input is the DECODED string (charset handling happens in the
    handler via the message_text charset primitives), so this is pure
    stdlib `csv` with no bytes-level surface. Reads rows until the
    accumulated output reaches `max_chars` (early-exit) or `_MAX_CSV_ROWS`
    rows are read. `csv.reader` handles quoting/embedded newlines; a
    malformed row degrades to ExtractionError rather than crashing.
    """
    try:
        out: list[str] = []
        total = 0
        reader = csv.reader(io.StringIO(text))
        for row_num, row in enumerate(reader):
            if row_num >= _MAX_CSV_ROWS or total >= max_chars:
                break
            line = "\t".join(row)
            out.append(line)
            total += len(line)
        return "\n".join(out)
    except RecursionError as exc:
        raise _safe("csv parse recursion limit") from exc
    except csv.Error as exc:
        raise _safe("csv parse error") from exc
    except Exception as exc:  # noqa: BLE001 - malformed CSV must not crash the read
        raise _safe("csv extraction failed") from exc
